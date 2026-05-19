"""Tests for the Office-file preview pipeline: convert_office_to_pdf +
the /api/chat/attachments/{id}/preview endpoint."""
from __future__ import annotations

import shutil
import struct
import zlib
from pathlib import Path

import pytest

from tokenmind.server.attachments import (
    MissingSofficeError,
    OfficeConversionError,
    convert_office_to_pdf,
    is_office_file,
)

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook  # noqa: E402


def _soffice_available() -> bool:
    if shutil.which("soffice"):
        return True
    return Path("/Applications/LibreOffice.app/Contents/MacOS/soffice").is_file()


# --- is_office_file ---------------------------------------------------------


def test_is_office_file_matches_known_extensions():
    assert is_office_file("report.xlsx")
    assert is_office_file("deck.pptx")
    assert is_office_file("contract.docx")
    assert is_office_file("/tmp/foo/old.xls")
    assert is_office_file(Path("readme.rtf"))
    assert is_office_file("plan.odt")


def test_is_office_file_rejects_other_formats():
    assert not is_office_file("photo.png")
    assert not is_office_file("notes.txt")
    assert not is_office_file("paper.pdf")
    assert not is_office_file("data.csv")  # CSV is plain text, not Office
    assert not is_office_file("")
    assert not is_office_file("no-extension")


def test_is_office_file_case_insensitive():
    assert is_office_file("REPORT.XLSX")
    assert is_office_file("Slide.PPTX")


# --- convert_office_to_pdf — error paths (no soffice needed) ----------------


def test_convert_missing_source_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        convert_office_to_pdf(tmp_path / "nope.xlsx")


def test_convert_missing_soffice_raises(tmp_path, monkeypatch):
    """When soffice is genuinely absent we surface MissingSofficeError, not
    a generic crash — the route uses that to return HTTP 503."""
    # Patch the resolver directly so we don't have to mock shutil + Path.is_file
    # (which fights pytest's own filesystem usage).
    def _missing() -> str:
        raise MissingSofficeError(
            "soffice (LibreOffice) is required. Install via 'brew install libreoffice'."
        )
    monkeypatch.setattr("tokenmind.server.attachments._find_soffice", _missing)

    src = tmp_path / "doc.xlsx"
    src.write_bytes(b"dummy")  # source must exist so we get past the FileNotFoundError check

    with pytest.raises(MissingSofficeError) as excinfo:
        convert_office_to_pdf(src)
    msg = str(excinfo.value).lower()
    assert "soffice" in msg or "libreoffice" in msg


# --- convert_office_to_pdf — happy path (needs real soffice) ---------------


@pytest.mark.skipif(not _soffice_available(), reason="soffice not installed")
def test_convert_xlsx_produces_pdf(tmp_path):
    src = tmp_path / "demo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Demo"
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    wb.save(src)

    pdf = convert_office_to_pdf(src)
    assert pdf.is_file()
    assert pdf.read_bytes().startswith(b"%PDF-")
    # Default cache lives next to source
    assert pdf == src.with_suffix(src.suffix + ".preview.pdf")


@pytest.mark.skipif(not _soffice_available(), reason="soffice not installed")
def test_convert_uses_cache_on_second_call(tmp_path):
    src = tmp_path / "cached.xlsx"
    Workbook().save(src)

    pdf1 = convert_office_to_pdf(src)
    mtime1 = pdf1.stat().st_mtime

    pdf2 = convert_office_to_pdf(src)
    assert pdf2 == pdf1
    # Same mtime → wasn't regenerated
    assert pdf2.stat().st_mtime == mtime1


@pytest.mark.skipif(not _soffice_available(), reason="soffice not installed")
def test_convert_regenerates_when_source_newer(tmp_path):
    src = tmp_path / "stale.xlsx"
    Workbook().save(src)

    pdf1 = convert_office_to_pdf(src)
    mtime_pdf = pdf1.stat().st_mtime

    # Touch source forward in time
    future = mtime_pdf + 5
    import os as _os
    _os.utime(src, (future, future))

    pdf2 = convert_office_to_pdf(src)
    assert pdf2 == pdf1
    assert pdf2.stat().st_mtime > mtime_pdf  # regenerated


@pytest.mark.skipif(not _soffice_available(), reason="soffice not installed")
def test_convert_respects_custom_cache_path(tmp_path):
    src = tmp_path / "wb.xlsx"
    Workbook().save(src)
    custom = tmp_path / "cache" / "out.pdf"

    pdf = convert_office_to_pdf(src, cache_path=custom)
    assert pdf == custom
    assert custom.is_file()


def test_convert_subprocess_nonzero_maps_to_conversion_error(tmp_path, monkeypatch):
    """If soffice exits non-zero, we surface OfficeConversionError so the
    route can return HTTP 502 instead of leaking a generic CalledProcessError."""
    src = tmp_path / "broken.xlsx"
    src.write_bytes(b"dummy")

    monkeypatch.setattr(
        "tokenmind.server.attachments._find_soffice",
        lambda: "/usr/bin/false-soffice-stub",
    )

    import subprocess as _sp

    class _FakeProc:
        def __init__(self):
            self.returncode = 1
            self.stdout = ""
            self.stderr = "Error: garbage input"

    def _run(*_args, **_kwargs):
        return _FakeProc()

    monkeypatch.setattr(_sp, "run", _run)

    with pytest.raises(OfficeConversionError) as excinfo:
        convert_office_to_pdf(src)
    assert "garbage input" in str(excinfo.value)


def test_convert_subprocess_timeout_maps_to_timeout_error(tmp_path, monkeypatch):
    """A subprocess.TimeoutExpired becomes our TimeoutError so the route
    returns HTTP 504 cleanly."""
    src = tmp_path / "slow.xlsx"
    src.write_bytes(b"dummy")

    monkeypatch.setattr(
        "tokenmind.server.attachments._find_soffice",
        lambda: "/usr/bin/false-soffice-stub",
    )

    import subprocess as _sp

    def _run(*_args, **_kwargs):
        raise _sp.TimeoutExpired(cmd="soffice", timeout=1)

    monkeypatch.setattr(_sp, "run", _run)

    with pytest.raises(TimeoutError):
        convert_office_to_pdf(src, timeout_s=1)


# --- /api/chat/attachments/{id}/preview endpoint ----------------------------


@pytest.fixture
def chat_service(monkeypatch, tmp_path):
    """Stub out the chat service so the FastAPI route uses a fake attachment store."""
    from fastapi.testclient import TestClient
    from fastapi import FastAPI
    from tokenmind.server.routes import chat as chat_module
    from tokenmind.server.dependencies import get_chat_service

    storage_path = tmp_path / "report.xlsx"
    wb = Workbook()
    wb.active["A1"] = "Hello"
    wb.save(storage_path)

    class FakeService:
        def __init__(self):
            self.attachments_db: dict = {
                "abc-xlsx": {
                    "id": "abc-xlsx",
                    "name": "report.xlsx",
                    "storage_path": str(storage_path),
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
                "abc-png": {
                    "id": "abc-png",
                    "name": "photo.png",
                    "storage_path": str(_make_tiny_png(tmp_path / "photo.png")),
                    "mime_type": "image/png",
                },
                "abc-missing-on-disk": {
                    "id": "abc-missing-on-disk",
                    "name": "ghost.docx",
                    "storage_path": str(tmp_path / "does-not-exist.docx"),
                    "mime_type": "application/msword",
                },
            }

        def resolve_attachment(self, attachment_id: str) -> dict:
            from fastapi import HTTPException
            if attachment_id not in self.attachments_db:
                raise HTTPException(status_code=404, detail="not found")
            return self.attachments_db[attachment_id]

    fake = FakeService()

    app = FastAPI()
    app.include_router(chat_module.router)
    # Override the LOCAL get_chat_service shim in chat.py (the route uses
    # ``Depends(get_chat_service)`` referencing the local symbol, not the
    # one in tokenmind.server.dependencies).
    app.dependency_overrides[chat_module.get_chat_service] = lambda: fake
    # Belt and suspenders: also override the canonical one in case some
    # path uses it directly.
    app.dependency_overrides[get_chat_service] = lambda: fake
    return TestClient(app), fake, storage_path


def _make_tiny_png(path: Path) -> Path:
    sig = b"\x89PNG\r\n\x1a\n"

    def chunk(tag: bytes, data: bytes) -> bytes:
        crc = zlib.crc32(tag + data) & 0xFFFFFFFF
        return struct.pack(">I", len(data)) + tag + data + struct.pack(">I", crc)

    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
    raw = b"\x00\xFF\x00\x00"
    idat = chunk(b"IDAT", zlib.compress(raw))
    iend = chunk(b"IEND", b"")
    path.write_bytes(sig + ihdr + idat + iend)
    return path


def test_preview_endpoint_serves_png_directly(chat_service):
    client, _, _ = chat_service
    r = client.get("/api/chat/attachments/abc-png/preview")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("image/")
    assert r.content.startswith(b"\x89PNG")


@pytest.mark.skipif(not _soffice_available(), reason="soffice not installed")
def test_preview_endpoint_converts_xlsx_to_pdf(chat_service):
    client, _, storage = chat_service
    r = client.get("/api/chat/attachments/abc-xlsx/preview")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content.startswith(b"%PDF-")
    # Cache file lives next to the xlsx
    assert (storage.with_suffix(storage.suffix + ".preview.pdf")).is_file()


def test_preview_endpoint_404_when_source_missing(chat_service):
    client, _, _ = chat_service
    r = client.get("/api/chat/attachments/abc-missing-on-disk/preview")
    assert r.status_code == 404


def test_preview_endpoint_unknown_id_404(chat_service):
    client, _, _ = chat_service
    r = client.get("/api/chat/attachments/does-not-exist/preview")
    assert r.status_code == 404


def test_preview_endpoint_503_when_soffice_missing(chat_service, monkeypatch):
    client, _, _ = chat_service
    # Force the conversion to behave as if soffice is unavailable
    def _missing(*_args, **_kwargs):
        raise MissingSofficeError("soffice not found, install LibreOffice")
    monkeypatch.setattr(
        "tokenmind.server.routes.chat.convert_office_to_pdf", _missing,
    )
    r = client.get("/api/chat/attachments/abc-xlsx/preview")
    assert r.status_code == 503
    assert "soffice" in r.json()["detail"].lower() or "libreoffice" in r.json()["detail"].lower()


def test_preview_endpoint_502_when_conversion_fails(chat_service, monkeypatch):
    client, _, _ = chat_service
    def _fail(*_args, **_kwargs):
        raise OfficeConversionError("soffice exited 1: garbage input")
    monkeypatch.setattr(
        "tokenmind.server.routes.chat.convert_office_to_pdf", _fail,
    )
    r = client.get("/api/chat/attachments/abc-xlsx/preview")
    assert r.status_code == 502


def test_preview_endpoint_504_on_timeout(chat_service, monkeypatch):
    client, _, _ = chat_service
    def _slow(*_args, **_kwargs):
        raise TimeoutError("soffice exceeded 60s")
    monkeypatch.setattr(
        "tokenmind.server.routes.chat.convert_office_to_pdf", _slow,
    )
    r = client.get("/api/chat/attachments/abc-xlsx/preview")
    assert r.status_code == 504
