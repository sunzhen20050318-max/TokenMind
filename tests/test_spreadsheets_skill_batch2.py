"""Tests for spreadsheets skill batch 2: add_chart / add_formula / format_cells."""
from __future__ import annotations

import importlib.util
import subprocess
import sys
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook  # noqa: E402

SKILL_DIR = Path(__file__).resolve().parent.parent / "tokenmind" / "skills" / "spreadsheets" / "scripts"


def _load_script(path: Path):
    spec = importlib.util.spec_from_file_location(path.stem, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[path.stem] = module
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def add_chart():
    return _load_script(SKILL_DIR / "add_chart.py")


@pytest.fixture(scope="module")
def add_formula():
    return _load_script(SKILL_DIR / "add_formula.py")


@pytest.fixture(scope="module")
def format_cells():
    return _load_script(SKILL_DIR / "format_cells.py")


def _make_sample_wb(tmp_path: Path) -> Path:
    """Workbook with a Sales sheet: A=Month, B=Revenue, 12 rows."""
    out = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    for i, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
        ws[f"A{i+2}"] = month
        ws[f"B{i+2}"] = (i + 1) * 100
    wb.save(out)
    return out


# --- add_chart ---------------------------------------------------------------


def test_add_chart_bar(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    total = add_chart.add_chart(
        workbook=wb_path,
        sheet="Sales",
        chart_type="bar",
        data_range="B1:B13",
        categories_range="A2:A13",
        anchor="D2",
        title="Revenue by Month",
        titles_from_data=True,
    )
    assert total == 1
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert len(ws._charts) == 1


def test_add_chart_line(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    add_chart.add_chart(
        workbook=wb_path,
        sheet="Sales",
        chart_type="line",
        data_range="B2:B13",
        categories_range="A2:A13",
        anchor="D2",
    )
    wb = load_workbook(wb_path)
    assert len(wb["Sales"]._charts) == 1


def test_add_chart_pie(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    add_chart.add_chart(
        workbook=wb_path,
        sheet="Sales",
        chart_type="pie",
        data_range="B2:B7",
        categories_range="A2:A7",
        anchor="D2",
    )
    wb = load_workbook(wb_path)
    assert len(wb["Sales"]._charts) == 1


def test_add_chart_invalid_type(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="--type"):
        add_chart.add_chart(
            workbook=wb_path, sheet="Sales", chart_type="donut",
            data_range="B1:B13", categories_range=None, anchor="D2",
        )


def test_add_chart_invalid_range(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="Invalid range"):
        add_chart.add_chart(
            workbook=wb_path, sheet="Sales", chart_type="bar",
            data_range="not-a-range", categories_range=None, anchor="D2",
        )


def test_add_chart_missing_sheet(tmp_path, add_chart):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="not found"):
        add_chart.add_chart(
            workbook=wb_path, sheet="Nope", chart_type="bar",
            data_range="B1:B13", categories_range=None, anchor="D2",
        )


# --- add_formula -------------------------------------------------------------


def test_add_formula_single_cell(tmp_path, add_formula):
    wb_path = _make_sample_wb(tmp_path)
    count = add_formula.set_formula(wb_path, "Sales", "C2", "=B2*1.1")
    assert count == 1
    wb = load_workbook(wb_path)
    assert wb["Sales"]["C2"].value == "=B2*1.1"


def test_add_formula_fill_down_row_token(tmp_path, add_formula):
    wb_path = _make_sample_wb(tmp_path)
    count = add_formula.fill_formula(wb_path, "Sales", "C2:C13", "=B{row}*1.2")
    assert count == 12
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert ws["C2"].value == "=B2*1.2"
    assert ws["C7"].value == "=B7*1.2"
    assert ws["C13"].value == "=B13*1.2"


def test_add_formula_fill_across_col_token(tmp_path, add_formula):
    wb_path = _make_sample_wb(tmp_path)
    count = add_formula.fill_formula(wb_path, "Sales", "B14:D14", "=SUM({col}2:{col}13)")
    assert count == 3
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert ws["B14"].value == "=SUM(B2:B13)"
    assert ws["C14"].value == "=SUM(C2:C13)"
    assert ws["D14"].value == "=SUM(D2:D13)"


def test_add_formula_missing_equals_rejected(tmp_path, add_formula):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="must start with"):
        add_formula.set_formula(wb_path, "Sales", "C2", "B2*1.1")


def test_add_formula_invalid_range_rejected(tmp_path, add_formula):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="Invalid range"):
        add_formula.fill_formula(wb_path, "Sales", "BADRANGE", "=1+1")


# --- format_cells ------------------------------------------------------------


def test_format_cells_bold_and_bg(tmp_path, format_cells):
    wb_path = _make_sample_wb(tmp_path)
    count = format_cells.format_cells(
        wb_path, "Sales", "A1:B1",
        bold=True, bg_color="1F2937", font_color="FFFFFF", align="center",
    )
    assert count == 2
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws["A1"].font.color.value == "00FFFFFF" or "FFFFFF" in str(ws["A1"].font.color.value or "")
    assert ws["A1"].fill.start_color.value == "001F2937" or "1F2937" in str(ws["A1"].fill.start_color.value or "")
    assert ws["A1"].alignment.horizontal == "center"


def test_format_cells_number_format(tmp_path, format_cells):
    wb_path = _make_sample_wb(tmp_path)
    format_cells.format_cells(wb_path, "Sales", "B2:B13", number_format="$#,##0.00")
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert ws["B2"].number_format == "$#,##0.00"
    assert ws["B13"].number_format == "$#,##0.00"


def test_format_cells_border(tmp_path, format_cells):
    wb_path = _make_sample_wb(tmp_path)
    format_cells.format_cells(wb_path, "Sales", "A1:B3", border="thin")
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert ws["A1"].border.left.style == "thin"
    assert ws["B3"].border.bottom.style == "thin"


def test_format_cells_bad_hex_rejected(tmp_path, format_cells):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="hex"):
        format_cells.format_cells(wb_path, "Sales", "A1:A1", bg_color="ZZZZZZ")


def test_format_cells_bad_align_rejected(tmp_path, format_cells):
    wb_path = _make_sample_wb(tmp_path)
    with pytest.raises(ValueError, match="--align"):
        format_cells.format_cells(wb_path, "Sales", "A1:A1", align="middle")


def test_format_cells_preserves_size_on_partial_bold(tmp_path, format_cells):
    """Setting only --bold shouldn't blow away the cell's existing font size."""
    wb_path = _make_sample_wb(tmp_path)
    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    ws["A1"].font = openpyxl.styles.Font(size=18)
    wb.save(wb_path)

    format_cells.format_cells(wb_path, "Sales", "A1:A1", bold=True)

    wb = load_workbook(wb_path)
    assert wb["Sales"]["A1"].font.size == 18
    assert wb["Sales"]["A1"].font.bold is True


# --- end-to-end CLI smoke ---------------------------------------------------


def test_cli_smoke_chart_formula_format(tmp_path):
    wb_path = _make_sample_wb(tmp_path)

    r = subprocess.run([
        sys.executable, str(SKILL_DIR / "add_chart.py"), str(wb_path),
        "--sheet", "Sales", "--type", "bar",
        "--data", "B1:B13", "--categories", "A2:A13",
        "--anchor", "D2", "--titles-from-data", "--title", "Revenue",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    r = subprocess.run([
        sys.executable, str(SKILL_DIR / "add_formula.py"), str(wb_path),
        "--sheet", "Sales", "--fill-range", "C2:C13",
        "--formula", "=B{row}*1.2",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    r = subprocess.run([
        sys.executable, str(SKILL_DIR / "format_cells.py"), str(wb_path),
        "--sheet", "Sales", "--range", "A1:C1",
        "--bold", "--bg-color", "1F2937", "--font-color", "FFFFFF", "--align", "center",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    wb = load_workbook(wb_path)
    ws = wb["Sales"]
    assert len(ws._charts) == 1
    assert ws["C2"].value == "=B2*1.2"
    assert ws["A1"].font.bold is True
