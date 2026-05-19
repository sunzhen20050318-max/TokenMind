"""Tests for spreadsheets skill batch 3: verify_xlsx / csv_to_xlsx."""
from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook  # noqa: E402

SKILL_DIR = Path(__file__).resolve().parent.parent / "tokenmind" / "skills" / "spreadsheets" / "scripts"


def _load_script(path: Path):
    spec = importlib.util.spec_from_file_location(path.stem, path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[path.stem] = module
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def verify_xlsx():
    return _load_script(SKILL_DIR / "verify_xlsx.py")


@pytest.fixture(scope="module")
def csv_to_xlsx():
    return _load_script(SKILL_DIR / "csv_to_xlsx.py")


def _make_wb(tmp_path: Path) -> Path:
    out = tmp_path / "wb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Header"
    ws["A2"] = "=SUM(B2:B4)"
    ws["B2"] = 10
    ws["B3"] = 20
    ws["B4"] = 30
    detail = wb.create_sheet("Detail")
    for i in range(1, 13):
        detail[f"A{i}"] = f"item{i}"
        detail[f"B{i}"] = i * 5
    empty = wb.create_sheet("Empty")
    _ = empty  # silence unused
    wb.save(out)
    return out


# --- verify_xlsx ------------------------------------------------------------


def test_verify_reports_basic_shape(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    assert report["sheet_count"] == 3
    assert report["sheet_names"] == ["Summary", "Detail", "Empty"]
    summary = next(s for s in report["sheets"] if s["title"] == "Summary")
    assert summary["max_row"] == 4
    assert summary["max_column"] == 2
    assert summary["formula_count"] == 1  # the SUM in A2
    detail = next(s for s in report["sheets"] if s["title"] == "Detail")
    assert detail["max_row"] == 12
    assert detail["cell_count"] == 24
    empty = next(s for s in report["sheets"] if s["title"] == "Empty")
    assert empty["cell_count"] == 0


def test_verify_expect_sheets_pass(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    failures = verify_xlsx.check_expectations(
        report, expect_sheets="Summary,Detail",
    )
    assert failures == []


def test_verify_expect_sheets_missing(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    failures = verify_xlsx.check_expectations(
        report, expect_sheets="Summary,Missing",
    )
    assert len(failures) == 1
    assert "Missing" in failures[0]


def test_verify_expect_min_rows(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    failures = verify_xlsx.check_expectations(
        report, expect_min_rows="Detail=12",
    )
    assert failures == []
    failures = verify_xlsx.check_expectations(
        report, expect_min_rows="Detail=100",
    )
    assert any("Detail" in f for f in failures)


def test_verify_expect_charts_expression(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    failures = verify_xlsx.check_expectations(
        report, expect_charts="Summary=0,Detail=0",
    )
    assert failures == []
    failures = verify_xlsx.check_expectations(
        report, expect_charts="Summary>=1",
    )
    assert any("chart_count" in f for f in failures)


def test_verify_expect_no_empty_sheets_catches_empty(tmp_path, verify_xlsx):
    wb_path = _make_wb(tmp_path)
    report = verify_xlsx.audit_workbook(wb_path)
    failures = verify_xlsx.check_expectations(
        report, expect_no_empty_sheets=True,
    )
    assert len(failures) == 1
    assert "Empty" in failures[0]


def test_verify_cli_json_output(tmp_path):
    wb_path = _make_wb(tmp_path)
    r = subprocess.run(
        [sys.executable, str(SKILL_DIR / "verify_xlsx.py"), str(wb_path)],
        capture_output=True, text=True,
    )
    assert r.returncode == 0, r.stderr
    payload = json.loads(r.stdout)
    assert payload["failures"] == []
    assert payload["report"]["sheet_count"] == 3


def test_verify_cli_nonzero_on_failure(tmp_path):
    wb_path = _make_wb(tmp_path)
    r = subprocess.run(
        [
            sys.executable,
            str(SKILL_DIR / "verify_xlsx.py"),
            str(wb_path),
            "--expect-charts",
            "Summary>=1",
        ],
        capture_output=True, text=True,
    )
    assert r.returncode == 1
    payload = json.loads(r.stdout)
    assert len(payload["failures"]) >= 1


# --- csv_to_xlsx -------------------------------------------------------------


def _make_csv(tmp_path: Path, name: str = "data.csv", rows: list[list[str]] | None = None) -> Path:
    rows = rows or [["Name", "Qty"], ["Widget", "10"], ["Gadget", "7"], ["Sprocket", "3"]]
    p = tmp_path / name
    p.write_text("\n".join(",".join(r) for r in rows) + "\n", encoding="utf-8")
    return p


def test_csv_to_xlsx_new_workbook(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    path, sheet, rows = csv_to_xlsx.csv_to_xlsx(csv_path, out=out)
    assert path == out
    assert sheet == "data"  # CSV stem
    assert rows == 4
    wb = load_workbook(out)
    assert wb.sheetnames == ["data"]
    ws = wb["data"]
    assert ws["A1"].value == "Name"
    assert ws["B2"].value == "10"  # stays as string by default


def test_csv_to_xlsx_existing_workbook(tmp_path, csv_to_xlsx):
    # Build a base wb first
    base = tmp_path / "base.xlsx"
    wb = Workbook()
    wb.active.title = "Cover"
    wb.save(base)

    csv_path = _make_csv(tmp_path)
    path, sheet, _ = csv_to_xlsx.csv_to_xlsx(csv_path, workbook=base, sheet="Imported")
    assert path == base
    assert sheet == "Imported"
    wb = load_workbook(base)
    assert wb.sheetnames == ["Cover", "Imported"]


def test_csv_to_xlsx_coerce_numbers(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    csv_to_xlsx.csv_to_xlsx(csv_path, out=out, coerce_numbers=True)
    wb = load_workbook(out)
    ws = wb["data"]
    assert ws["B2"].value == 10  # int now
    assert isinstance(ws["B2"].value, int)


def test_csv_to_xlsx_header_styling(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    csv_to_xlsx.csv_to_xlsx(csv_path, out=out, header_row=True)
    wb = load_workbook(out)
    ws = wb["data"]
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert ws["A2"].font.bold is not True  # data rows untouched


def test_csv_to_xlsx_refuses_clobber_without_overwrite(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    csv_to_xlsx.csv_to_xlsx(csv_path, out=out)
    with pytest.raises(FileExistsError):
        csv_to_xlsx.csv_to_xlsx(csv_path, out=out)


def test_csv_to_xlsx_overwrite_flag(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    csv_to_xlsx.csv_to_xlsx(csv_path, out=out)
    csv_to_xlsx.csv_to_xlsx(csv_path, out=out, overwrite=True)
    # Just confirm it didn't raise


def test_csv_to_xlsx_either_out_or_workbook(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    with pytest.raises(ValueError, match="exactly one"):
        csv_to_xlsx.csv_to_xlsx(csv_path)  # neither
    with pytest.raises(ValueError, match="exactly one"):
        csv_to_xlsx.csv_to_xlsx(csv_path, out=tmp_path / "a.xlsx", workbook=tmp_path / "b.xlsx")


def test_csv_to_xlsx_long_sheet_name_truncated(tmp_path, csv_to_xlsx):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    long_name = "x" * 50
    _, sheet, _ = csv_to_xlsx.csv_to_xlsx(csv_path, out=out, sheet=long_name)
    assert len(sheet) == 31  # Excel hard cap


def test_csv_to_xlsx_duplicate_sheet_name_in_existing_rejected(tmp_path, csv_to_xlsx):
    base = tmp_path / "base.xlsx"
    wb = Workbook()
    wb.active.title = "Imported"
    wb.save(base)

    csv_path = _make_csv(tmp_path)
    with pytest.raises(ValueError, match="already exists"):
        csv_to_xlsx.csv_to_xlsx(csv_path, workbook=base, sheet="Imported")


def test_csv_to_xlsx_cli_smoke(tmp_path):
    csv_path = _make_csv(tmp_path)
    out = tmp_path / "out.xlsx"
    r = subprocess.run(
        [
            sys.executable,
            str(SKILL_DIR / "csv_to_xlsx.py"),
            str(csv_path),
            "--out", str(out),
            "--header-row",
            "--auto-fit",
            "--coerce-numbers",
        ],
        capture_output=True, text=True,
    )
    assert r.returncode == 0, r.stderr
    wb = load_workbook(out)
    ws = wb["data"]
    assert ws["B2"].value == 10
    assert ws["A1"].font.bold is True
