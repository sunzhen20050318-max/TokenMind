#!/usr/bin/env python3
"""Audit a workbook's structure and optionally assert expectations.

By default this script just reports — exit 0 with a JSON shape summary
on stdout. Pass any of the ``--expect-*`` flags to turn it into a gate
(non-zero exit + detailed failure list on stderr).

Examples
--------

  # Plain structural report (JSON)
  python verify_xlsx.py /tmp/wb.xlsx

  # Human-readable report
  python verify_xlsx.py /tmp/wb.xlsx --format text

  # Assert specific sheets exist (order doesn't matter)
  python verify_xlsx.py /tmp/wb.xlsx \\
      --expect-sheets "Summary,Detail,Assumptions"

  # Assert minimum data shape per sheet
  python verify_xlsx.py /tmp/wb.xlsx \\
      --expect-min-rows "Detail=12,Summary=4"

  # Assert charts present
  python verify_xlsx.py /tmp/wb.xlsx --expect-charts "Summary>=1"

Reported fields per sheet
-------------------------

  - title
  - max_row, max_column (openpyxl's used-range bounds)
  - cell_count (non-empty cells)
  - formula_count
  - chart_count
  - merged_ranges (count)
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


_EXPECT_RE = re.compile(r"^([^=<>]+?)\s*(>=|<=|=|>|<)\s*(\d+)\s*$")


def _audit_sheet(ws) -> dict[str, Any]:
    cell_count = 0
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell_count += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_count += 1
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "cell_count": cell_count,
        "formula_count": formula_count,
        "chart_count": len(ws._charts),  # type: ignore[attr-defined]
        "merged_ranges": len(ws.merged_cells.ranges),
    }


def audit_workbook(workbook: Path) -> dict[str, Any]:
    wb = load_workbook(workbook)
    sheets = [_audit_sheet(wb[name]) for name in wb.sheetnames]
    return {
        "path": str(workbook),
        "sheet_count": len(sheets),
        "sheet_names": [s["title"] for s in sheets],
        "sheets": sheets,
    }


def _parse_kv_pairs(arg: str) -> list[tuple[str, int]]:
    """Parse ``Sheet=12,Other=3`` → [('Sheet', 12), ('Other', 3)]."""
    out = []
    for piece in arg.split(","):
        piece = piece.strip()
        if not piece:
            continue
        if "=" not in piece:
            raise ValueError(f"Expected 'Sheet=N', got {piece!r}")
        name, val = piece.split("=", 1)
        out.append((name.strip(), int(val.strip())))
    return out


def _parse_expression(expr: str) -> tuple[str, str, int]:
    """Parse ``Summary>=1`` → ('Summary', '>=', 1)."""
    m = _EXPECT_RE.match(expr)
    if not m:
        raise ValueError(f"Expected 'Name<op>N', got {expr!r}")
    return m.group(1).strip(), m.group(2), int(m.group(3))


def _cmp(actual: int, op: str, expected: int) -> bool:
    return {
        ">=": actual >= expected,
        "<=": actual <= expected,
        "=":  actual == expected,
        ">":  actual >  expected,
        "<":  actual <  expected,
    }[op]


def check_expectations(
    report: dict[str, Any],
    *,
    expect_sheets: str | None = None,
    expect_min_rows: str | None = None,
    expect_charts: str | None = None,
    expect_no_empty_sheets: bool = False,
) -> list[str]:
    """Return a list of failure messages (empty list = all good)."""
    failures: list[str] = []
    by_name = {s["title"]: s for s in report["sheets"]}

    if expect_sheets:
        wanted = {s.strip() for s in expect_sheets.split(",") if s.strip()}
        actual = set(report["sheet_names"])
        missing = wanted - actual
        if missing:
            failures.append(f"missing sheets: {sorted(missing)}")

    if expect_min_rows:
        for sheet, min_rows in _parse_kv_pairs(expect_min_rows):
            if sheet not in by_name:
                failures.append(f"--expect-min-rows references unknown sheet {sheet!r}")
                continue
            if by_name[sheet]["max_row"] < min_rows:
                failures.append(
                    f"sheet {sheet!r}: max_row={by_name[sheet]['max_row']} "
                    f"< expected {min_rows}"
                )

    if expect_charts:
        for piece in expect_charts.split(","):
            piece = piece.strip()
            if not piece:
                continue
            sheet, op, expected = _parse_expression(piece)
            if sheet not in by_name:
                failures.append(f"--expect-charts references unknown sheet {sheet!r}")
                continue
            actual = by_name[sheet]["chart_count"]
            if not _cmp(actual, op, expected):
                failures.append(
                    f"sheet {sheet!r}: chart_count={actual} fails {op}{expected}"
                )

    if expect_no_empty_sheets:
        for s in report["sheets"]:
            if s["cell_count"] == 0:
                failures.append(f"sheet {s['title']!r} is empty")

    return failures


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--format", choices=["json", "text"], default="json", help="Output format (default: json)")
    parser.add_argument("--expect-sheets", default=None, help="Comma-separated sheet names that must exist")
    parser.add_argument(
        "--expect-min-rows",
        default=None,
        help="Comma-separated 'Sheet=N' minimum max_row per sheet",
    )
    parser.add_argument(
        "--expect-charts",
        default=None,
        help="Comma-separated 'Sheet<op>N' chart-count expressions, e.g. 'Summary>=1,Detail=0'",
    )
    parser.add_argument(
        "--expect-no-empty-sheets",
        action="store_true",
        help="Fail if any sheet has zero non-empty cells",
    )
    args = parser.parse_args(argv)

    if not args.workbook.is_file():
        print(f"Error: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        report = audit_workbook(args.workbook)
    except Exception as e:  # noqa: BLE001
        print(f"Error: failed to load workbook: {e}", file=sys.stderr)
        return 1

    failures = check_expectations(
        report,
        expect_sheets=args.expect_sheets,
        expect_min_rows=args.expect_min_rows,
        expect_charts=args.expect_charts,
        expect_no_empty_sheets=args.expect_no_empty_sheets,
    )

    if args.format == "json":
        print(json.dumps({"report": report, "failures": failures}, indent=2, ensure_ascii=False))
    else:
        print(f"Workbook: {report['path']}")
        print(f"  sheets: {report['sheet_count']}")
        for s in report["sheets"]:
            print(
                f"  - {s['title']}: max_row={s['max_row']} max_col={s['max_column']} "
                f"cells={s['cell_count']} formulas={s['formula_count']} "
                f"charts={s['chart_count']} merged={s['merged_ranges']}"
            )
        if failures:
            print("Failures:")
            for f in failures:
                print(f"  - {f}")
        else:
            print("OK")

    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
