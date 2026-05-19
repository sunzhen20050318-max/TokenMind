#!/usr/bin/env python3
"""Apply formatting to a cell range: font, fill, alignment, number format,
borders.

Examples
--------

  # Header row: bold + dark fill + white text + centered
  python format_cells.py /tmp/wb.xlsx --sheet Main --range A1:F1 \\
      --bold --bg-color 1F2937 --font-color FFFFFF --align center

  # Currency column
  python format_cells.py /tmp/wb.xlsx --sheet Main --range D2:D100 \\
      --number-format "$#,##0.00"

  # Percentage
  python format_cells.py /tmp/wb.xlsx --sheet Main --range E2:E100 \\
      --number-format "0.00%"

  # Outline border around a range
  python format_cells.py /tmp/wb.xlsx --sheet Main --range A1:F10 \\
      --border thin

Hex colors are 6-digit RGB without leading ``#`` (openpyxl convention).
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import column_index_from_string, get_column_letter
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


_HEX_RE = re.compile(r"^[0-9A-Fa-f]{6}$")
_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE)


def _check_hex(value: str | None, label: str) -> str | None:
    if value is None:
        return None
    v = value.lstrip("#")
    if not _HEX_RE.match(v):
        raise ValueError(f"{label} must be 6-digit hex (e.g. 1F2937), got {value!r}")
    return v.upper()


def _iter_cells(ws, rng: str):
    m = _RANGE_RE.match(rng.strip())
    if not m:
        raise ValueError(f"Invalid range {rng!r}. Expected form like 'A1:B12'.")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    min_col = column_index_from_string(c1)
    max_col = column_index_from_string(c2)
    min_row, max_row = min(r1, r2), max(r1, r2)
    if min_col > max_col:
        min_col, max_col = max_col, min_col
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            yield ws[f"{get_column_letter(col)}{row}"]


def format_cells(
    workbook: Path,
    sheet: str,
    range_: str,
    *,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    align: str | None = None,
    wrap: bool = False,
    number_format: str | None = None,
    border: str | None = None,
) -> int:
    """Apply formatting flags to every cell in ``range_``. Returns cell count."""
    if align is not None and align not in {"left", "center", "right"}:
        raise ValueError(f"--align must be left|center|right, got {align!r}")
    if border is not None and border not in {"thin", "medium", "thick"}:
        raise ValueError(f"--border must be thin|medium|thick, got {border!r}")

    fc = _check_hex(font_color, "--font-color")
    bg = _check_hex(bg_color, "--bg-color")

    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    side = Side(style=border) if border else None
    border_obj = Border(left=side, right=side, top=side, bottom=side) if side else None
    fill_obj = PatternFill(start_color=bg, end_color=bg, fill_type="solid") if bg else None

    count = 0
    for cell in _iter_cells(ws, range_):
        # Merge font attrs with the cell's existing font so we don't clobber size
        # on a partial bold/italic update.
        existing = cell.font
        cell.font = Font(
            name=existing.name,
            size=font_size if font_size is not None else existing.size,
            bold=bold or existing.bold,
            italic=italic or existing.italic,
            underline="single" if underline else existing.underline,
            color=fc if fc is not None else existing.color,
        )
        if fill_obj is not None:
            cell.fill = fill_obj
        if align or wrap:
            cell.alignment = Alignment(
                horizontal=align,
                vertical=cell.alignment.vertical,
                wrap_text=wrap or cell.alignment.wrap_text,
            )
        if number_format:
            cell.number_format = number_format
        if border_obj is not None:
            cell.border = border_obj
        count += 1

    wb.save(workbook)
    return count


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--sheet", required=True, help="Target sheet name")
    parser.add_argument("--range", dest="range_", required=True, help="Range to format, e.g. A1:F1")

    parser.add_argument("--bold", action="store_true")
    parser.add_argument("--italic", action="store_true")
    parser.add_argument("--underline", action="store_true")
    parser.add_argument("--font-size", type=int, default=None)
    parser.add_argument("--font-color", default=None, help="6-digit hex like FFFFFF (no '#')")
    parser.add_argument("--bg-color", default=None, help="6-digit hex like 1F2937 (no '#')")
    parser.add_argument("--align", choices=["left", "center", "right"], default=None)
    parser.add_argument("--wrap", action="store_true", help="Enable text wrap in cells")
    parser.add_argument("--number-format", dest="number_format", default=None, help='e.g. "$#,##0.00" or "0.00%"')
    parser.add_argument("--border", choices=["thin", "medium", "thick"], default=None, help="Outline border style")

    args = parser.parse_args(argv)
    if not args.workbook.is_file():
        print(f"Error: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        count = format_cells(
            workbook=args.workbook,
            sheet=args.sheet,
            range_=args.range_,
            bold=args.bold,
            italic=args.italic,
            underline=args.underline,
            font_size=args.font_size,
            font_color=args.font_color,
            bg_color=args.bg_color,
            align=args.align,
            wrap=args.wrap,
            number_format=args.number_format,
            border=args.border,
        )
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Formatted {count} cell(s) in {args.workbook} [{args.sheet}!{args.range_}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
