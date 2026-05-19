#!/usr/bin/env python3
"""Write cell values into an existing .xlsx workbook (in place).

Supports three input shapes via mutually-exclusive flags:

1. ``--cells JSON`` — single-cell map. JSON object keyed by cell ref:
   ``{"A1": "Title", "B2": 42, "C3": "=SUM(B2:B5)"}``

2. ``--rows JSON --start A1`` — dense 2D grid. JSON array of rows, each row
   an array of cell values, written starting at ``--start`` (default A1):
   ``[["Name","Qty"],["Widget",10],["Gadget",7]]``

3. ``--csv FILE --start A1`` — load rows from a CSV file, written starting
   at ``--start``. The first row is treated as data, not as a header.

A leading ``=`` in a string value is preserved as a formula
(``"=SUM(...)"``). Numbers/booleans are passed through as native types.

Examples
--------

  python set_values.py /tmp/wb.xlsx --sheet Sheet1 \\
      --cells '{"A1":"Revenue","B1":1200}'

  python set_values.py /tmp/wb.xlsx --sheet Detail \\
      --rows '[["Name","Qty"],["Widget",10]]' --start A1

  python set_values.py /tmp/wb.xlsx --sheet Import \\
      --csv ./data.csv --start A2
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


def _parse_value(v: Any) -> Any:
    """openpyxl writes strings starting with '=' as formulas already, so we
    just return the value unchanged. This helper is here so future coercion
    (date parsing, percent strings, etc.) has a single chokepoint."""
    return v


def set_cells(workbook: Path, sheet: str, cells: dict[str, Any]) -> int:
    """Set values from a {cell_ref: value} map. Returns count written."""
    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    for ref, val in cells.items():
        # Validate cell ref by parsing it — raises on bad input
        coordinate_from_string(ref)
        ws[ref] = _parse_value(val)
    wb.save(workbook)
    return len(cells)


def set_rows(
    workbook: Path,
    sheet: str,
    rows: list[list[Any]],
    start: str = "A1",
) -> int:
    """Write a 2D ``rows`` array starting at cell ``start``. Returns cell count."""
    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    col_letter, row_idx = coordinate_from_string(start)
    start_col = column_index_from_string(col_letter)
    count = 0
    for r_offset, row in enumerate(rows):
        for c_offset, val in enumerate(row):
            ref = f"{get_column_letter(start_col + c_offset)}{row_idx + r_offset}"
            ws[ref] = _parse_value(val)
            count += 1
    wb.save(workbook)
    return count


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--sheet", required=True, help="Target sheet name")
    parser.add_argument("--start", default="A1", help="Origin cell for --rows/--csv (default: A1)")

    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--cells", help="JSON object: {\"A1\": value, ...}")
    src.add_argument("--rows", help="JSON 2D array: [[r1c1, r1c2], [r2c1, r2c2], ...]")
    src.add_argument("--csv", type=Path, help="Path to a CSV file to import")

    args = parser.parse_args(argv)

    if not args.workbook.is_file():
        print(f"Error: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        if args.cells:
            spec = json.loads(args.cells)
            if not isinstance(spec, dict):
                raise ValueError("--cells must be a JSON object")
            count = set_cells(args.workbook, args.sheet, spec)
        elif args.rows:
            spec = json.loads(args.rows)
            if not isinstance(spec, list):
                raise ValueError("--rows must be a JSON array")
            count = set_rows(args.workbook, args.sheet, spec, args.start)
        else:
            with args.csv.open("r", newline="", encoding="utf-8") as fh:
                rows = list(csv.reader(fh))
            count = set_rows(args.workbook, args.sheet, rows, args.start)
    except (ValueError, json.JSONDecodeError) as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Wrote {count} cell(s) to {args.workbook} [{args.sheet}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
