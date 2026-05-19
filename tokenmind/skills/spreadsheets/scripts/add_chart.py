#!/usr/bin/env python3
"""Add a chart (bar / line / pie) to a sheet in an .xlsx workbook.

Examples
--------

  # Bar chart with categories in column A and one data series in column B
  python add_chart.py /tmp/wb.xlsx --sheet Sales \\
      --type bar --data B1:B12 --categories A2:A12 \\
      --titles-from-data --anchor D2 --title "Monthly Revenue"

  # Line chart, multiple data series (each column is its own series)
  python add_chart.py /tmp/wb.xlsx --sheet KPI \\
      --type line --data B1:D12 --categories A2:A12 \\
      --titles-from-data --anchor F2

  # Pie chart — first column data, first column categories
  python add_chart.py /tmp/wb.xlsx --sheet Mix \\
      --type pie --data B2:B7 --categories A2:A7 --anchor D2

Range syntax
------------

  ``--data`` accepts any valid range, e.g. ``B1:B12`` or ``B1:D12``.
  Include the **header row** when passing ``--titles-from-data`` so the
  first row of each column becomes the series legend.

  ``--categories`` should be a single column or row of category labels,
  *excluding* the header (or include it consistently — your call).
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


_CHART_CLASSES = {"bar": BarChart, "line": LineChart, "pie": PieChart}
_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE)


def _parse_range(rng: str) -> tuple[int, int, int, int]:
    """Parse ``A1:B12`` → (min_col, min_row, max_col, max_row), all 1-based."""
    m = _RANGE_RE.match(rng.strip())
    if not m:
        raise ValueError(f"Invalid range {rng!r}. Expected form like 'A1:B12'.")
    col1, row1, col2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    return (
        column_index_from_string(col1),
        row1,
        column_index_from_string(col2),
        row2,
    )


def add_chart(
    workbook: Path,
    sheet: str,
    chart_type: str,
    data_range: str,
    categories_range: str | None,
    anchor: str,
    title: str | None = None,
    titles_from_data: bool = False,
) -> int:
    """Add one chart to ``sheet``. Returns the new chart count on the sheet."""
    if chart_type not in _CHART_CLASSES:
        raise ValueError(
            f"--type must be one of {sorted(_CHART_CLASSES)}, got {chart_type!r}"
        )
    # Validate the anchor cell — openpyxl will accept malformed strings and
    # silently put the chart somewhere weird otherwise.
    coordinate_from_string(anchor)

    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    chart = _CHART_CLASSES[chart_type]()
    if title:
        chart.title = title

    min_col, min_row, max_col, max_row = _parse_range(data_range)
    data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=titles_from_data)

    if categories_range:
        c_min_col, c_min_row, c_max_col, c_max_row = _parse_range(categories_range)
        cats_ref = Reference(ws, min_col=c_min_col, min_row=c_min_row, max_col=c_max_col, max_row=c_max_row)
        chart.set_categories(cats_ref)

    ws.add_chart(chart, anchor)
    wb.save(workbook)
    return len(ws._charts)  # type: ignore[attr-defined]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--sheet", required=True, help="Sheet containing the source data and where the chart anchors")
    parser.add_argument("--type", required=True, choices=sorted(_CHART_CLASSES), help="Chart type")
    parser.add_argument("--data", required=True, help="Data range, e.g. B1:B12 or B1:D12")
    parser.add_argument("--categories", default=None, help="Category labels range (single column or row)")
    parser.add_argument("--anchor", default="D2", help="Top-left cell where the chart is placed (default: D2)")
    parser.add_argument("--title", default=None, help="Chart title")
    parser.add_argument(
        "--titles-from-data",
        action="store_true",
        help="Treat the first row of --data as series titles (legend labels)",
    )
    args = parser.parse_args(argv)

    if not args.workbook.is_file():
        print(f"Error: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        total = add_chart(
            workbook=args.workbook,
            sheet=args.sheet,
            chart_type=args.type,
            data_range=args.data,
            categories_range=args.categories,
            anchor=args.anchor,
            title=args.title,
            titles_from_data=args.titles_from_data,
        )
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Added {args.type} chart at {args.anchor} on {args.sheet!r} (total charts on sheet: {total})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
