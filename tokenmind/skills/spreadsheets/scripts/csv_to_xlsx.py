#!/usr/bin/env python3
"""Import a CSV file into a workbook as a new (or existing) sheet.

This is sugar over ``build_xlsx + add_sheet + set_values`` for the common
case of "I have a CSV, I want it as a clean sheet".

Examples
--------

  # New workbook, sheet defaults to the CSV stem ("sales.csv" → "sales")
  python csv_to_xlsx.py ./sales.csv --out /tmp/wb.xlsx

  # Into existing workbook, explicit sheet name, header row gets bold styling
  python csv_to_xlsx.py ./data.csv --workbook /tmp/wb.xlsx \\
      --sheet "Imported" --header-row --auto-fit

  # Coerce numeric-looking strings into numbers (default: leave as strings)
  python csv_to_xlsx.py ./prices.csv --out /tmp/wb.xlsx --coerce-numbers

  # Custom delimiter
  python csv_to_xlsx.py ./pipe.psv --out /tmp/wb.xlsx --delimiter "|"

Either ``--out`` or ``--workbook`` is required:
  * ``--out PATH``  → create a new workbook (errors if PATH exists,
    unless ``--overwrite`` is set)
  * ``--workbook PATH`` → append to an existing workbook
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


def _maybe_number(value: str) -> Any:
    """Try int → float → original string."""
    s = value.strip()
    if s == "":
        return value
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return value


def csv_to_xlsx(
    csv_path: Path,
    *,
    out: Path | None = None,
    workbook: Path | None = None,
    sheet: str | None = None,
    delimiter: str = ",",
    header_row: bool = False,
    auto_fit: bool = False,
    coerce_numbers: bool = False,
    overwrite: bool = False,
) -> tuple[Path, str, int]:
    """Import csv → xlsx. Returns ``(output_path, sheet_name, rows_written)``.

    Exactly one of ``out`` / ``workbook`` must be set.
    """
    if (out is None) == (workbook is None):
        raise ValueError("Specify exactly one of --out or --workbook")

    if not csv_path.is_file():
        raise FileNotFoundError(csv_path)

    sheet_name = sheet or csv_path.stem
    if len(sheet_name) > 31:  # Excel hard cap
        sheet_name = sheet_name[:31]

    if out is not None:
        if out.exists() and not overwrite:
            raise FileExistsError(
                f"{out} already exists. Pass --overwrite to replace it."
            )
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        # Replace the auto-created sheet so we don't leave "Sheet1" behind
        default = wb.active
        if default is not None:
            wb.remove(default)
        ws = wb.create_sheet(title=sheet_name)
        target_path = out
    else:
        if not workbook.is_file():
            raise FileNotFoundError(workbook)
        wb = load_workbook(workbook)
        if sheet_name in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} already exists in {workbook}. "
                f"Pick a different --sheet name."
            )
        ws = wb.create_sheet(title=sheet_name)
        target_path = workbook

    rows_written = 0
    max_widths: dict[int, int] = {}
    with csv_path.open("r", newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, val in enumerate(row, start=1):
                if coerce_numbers:
                    val = _maybe_number(val)
                ws.cell(row=r_idx, column=c_idx, value=val)
                # Track column width if auto-fitting
                if auto_fit:
                    width = len(str(val))
                    if width > max_widths.get(c_idx, 0):
                        max_widths[c_idx] = width
            rows_written += 1

    if header_row and rows_written > 0:
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

    if auto_fit:
        # openpyxl can't measure pixel widths reliably without a render
        # backend, so we use a char-count heuristic with sensible bounds.
        for c_idx, max_chars in max_widths.items():
            target = min(max(max_chars + 2, 6), 60)
            ws.column_dimensions[get_column_letter(c_idx)].width = target

    wb.save(target_path)
    return target_path, sheet_name, rows_written


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("csv", type=Path, help="Source CSV file")

    out_group = parser.add_mutually_exclusive_group(required=True)
    out_group.add_argument("--out", type=Path, help="Path to NEW workbook")
    out_group.add_argument("--workbook", type=Path, help="Path to EXISTING workbook")

    parser.add_argument("--sheet", default=None, help="Sheet name (default: CSV file stem)")
    parser.add_argument("--delimiter", default=",", help="CSV delimiter (default: comma)")
    parser.add_argument("--header-row", action="store_true", help="Style first row as bold header")
    parser.add_argument("--auto-fit", action="store_true", help="Heuristic column auto-fit")
    parser.add_argument("--coerce-numbers", action="store_true", help="Convert numeric strings to int/float")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite --out if it exists")
    args = parser.parse_args(argv)

    try:
        path, sheet_name, count = csv_to_xlsx(
            csv_path=args.csv,
            out=args.out,
            workbook=args.workbook,
            sheet=args.sheet,
            delimiter=args.delimiter,
            header_row=args.header_row,
            auto_fit=args.auto_fit,
            coerce_numbers=args.coerce_numbers,
            overwrite=args.overwrite,
        )
    except (FileNotFoundError, FileExistsError, ValueError) as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Imported {count} row(s) from {args.csv} → {path} [{sheet_name}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
