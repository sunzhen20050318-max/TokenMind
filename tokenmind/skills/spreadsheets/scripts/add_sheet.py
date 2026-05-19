#!/usr/bin/env python3
"""Add a new sheet to an existing .xlsx workbook (in place).

Examples
--------

  # Append a sheet named "Assumptions"
  python add_sheet.py /tmp/wb.xlsx --name Assumptions

  # Insert as the first tab
  python add_sheet.py /tmp/wb.xlsx --name Cover --position 0

  # Be tolerant — succeed if a sheet with this name already exists
  python add_sheet.py /tmp/wb.xlsx --name Detail --if-exists ignore
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


def add_sheet(
    workbook: Path,
    name: str,
    position: int | None = None,
    if_exists: str = "error",
) -> str:
    """Add a sheet named ``name`` to ``workbook``. Returns the final sheet name."""
    if if_exists not in {"error", "ignore", "rename"}:
        raise ValueError(f"--if-exists must be one of error|ignore|rename, got {if_exists!r}")

    if not workbook.is_file():
        raise FileNotFoundError(workbook)

    wb = load_workbook(workbook)
    if name in wb.sheetnames:
        if if_exists == "error":
            raise ValueError(f"Sheet {name!r} already exists in {workbook}")
        if if_exists == "ignore":
            return name
        # rename: pick a unique suffix
        i = 2
        candidate = f"{name} ({i})"
        while candidate in wb.sheetnames:
            i += 1
            candidate = f"{name} ({i})"
        name = candidate

    ws = wb.create_sheet(title=name, index=position)
    wb.save(workbook)
    return ws.title


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--name", required=True, help="Name of the new sheet")
    parser.add_argument(
        "--position",
        type=int,
        default=None,
        help="Zero-based insertion index (default: append at end)",
    )
    parser.add_argument(
        "--if-exists",
        choices=["error", "ignore", "rename"],
        default="error",
        help="Behavior when a sheet with this name already exists",
    )
    args = parser.parse_args(argv)

    try:
        final = add_sheet(args.workbook, args.name, args.position, args.if_exists)
    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    print(f"Added sheet {final!r} to {args.workbook}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
