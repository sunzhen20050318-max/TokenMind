#!/usr/bin/env python3
"""Write a formula into one cell, or fill it down/across a range with row /
column substitution.

Examples
--------

  # Single cell
  python add_formula.py /tmp/wb.xlsx --sheet Main \\
      --cell D2 --formula "=SUM(A2:C2)"

  # Fill down — {row} is substituted per row
  python add_formula.py /tmp/wb.xlsx --sheet Main \\
      --fill-range D2:D11 --formula "=SUM(A{row}:C{row})"

  # Fill across — {col} is the column letter for each cell
  python add_formula.py /tmp/wb.xlsx --sheet Main \\
      --fill-range B12:E12 --formula "=SUM({col}2:{col}11)"

Template tokens
---------------

  ``{row}`` → current row number (e.g. 2, 3, 4, ...)
  ``{col}`` → current column letter (e.g. ``A``, ``B``, ``C``, ...)

  Both work in ``--fill-range`` mode. ``{row}`` works in a single-cell
  ``--cell`` write too (rare but useful for ``{row}=1`` literals).
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:  # pragma: no cover - guarded by skill requires.python
    print(
        "Error: openpyxl is not installed. Run: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)


_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", re.IGNORECASE)


def _expand(formula: str, row: int, col: int) -> str:
    """Substitute ``{row}`` / ``{col}`` placeholders in ``formula``."""
    return formula.replace("{row}", str(row)).replace("{col}", get_column_letter(col))


def set_formula(workbook: Path, sheet: str, cell: str, formula: str) -> int:
    if not formula.startswith("="):
        raise ValueError(f"Formula must start with '=', got {formula!r}")
    coordinate_from_string(cell)  # validate ref
    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    col_letter, row = coordinate_from_string(cell)
    col = column_index_from_string(col_letter)
    ws[cell] = _expand(formula, row, col)
    wb.save(workbook)
    return 1


def fill_formula(workbook: Path, sheet: str, fill_range: str, formula: str) -> int:
    """Fill ``fill_range`` with ``formula``, substituting {row} / {col} per cell."""
    if not formula.startswith("="):
        raise ValueError(f"Formula must start with '=', got {formula!r}")

    m = _RANGE_RE.match(fill_range.strip())
    if not m:
        raise ValueError(f"Invalid range {fill_range!r}. Expected form like 'A1:B12'.")
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    min_col = column_index_from_string(c1)
    max_col = column_index_from_string(c2)
    min_row, max_row = min(r1, r2), max(r1, r2)
    if min_col > max_col:
        min_col, max_col = max_col, min_col

    wb = load_workbook(workbook)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    count = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ref = f"{get_column_letter(col)}{row}"
            ws[ref] = _expand(formula, row, col)
            count += 1
    wb.save(workbook)
    return count


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx file")
    parser.add_argument("--sheet", required=True, help="Target sheet name")
    parser.add_argument("--formula", required=True, help="Formula starting with =. Supports {row} / {col} templates.")

    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--cell", help="Single cell ref, e.g. D2")
    target.add_argument("--fill-range", dest="fill_range", help="Range, e.g. D2:D11 (formula gets {row}/{col} substituted per cell)")

    args = parser.parse_args(argv)

    if not args.workbook.is_file():
        print(f"Error: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        if args.cell:
            count = set_formula(args.workbook, args.sheet, args.cell, args.formula)
        else:
            count = fill_formula(args.workbook, args.sheet, args.fill_range, args.formula)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1

    print(f"Wrote {count} formula cell(s) to {args.workbook} [{args.sheet}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
